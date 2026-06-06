import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
import os

def to_excel(all_records, filename):
    if not all_records:
        print("No data to export.")
        return
        
    df = pd.DataFrame(all_records)
    
    if df.empty:
        print("DataFrame is empty.")
        return
        
    # Prepare Daily Detail DataFrame
    df_daily = df.copy()
    df_daily.insert(0, '#', range(1, 1 + len(df_daily)))
    
    daily_columns = ['#', 'assetCode', 'date', 'runTime', 'downtime', 'downtimePct', 'idlePct']
    df_daily = df_daily[daily_columns]
    df_daily.columns = ['#', 'Asset Code', 'Date', 'Run Time (h)', 'Downtime (h)', 'Downtime %', 'Idle %']
    
    # Prepare Weekly Summary DataFrame
    summary_data = []
    for asset, group in df.groupby('assetCode'):
        valid_group = group[group['totalTime'] > 0]
        
        if valid_group.empty:
            avg_util = 0.0
            avg_down_pct = 0.0
            max_down_pct = 0.0
        else:
            avg_util = valid_group['utilization'].mean() * 100
            avg_down_pct = valid_group['downtimePct'].mean()
            max_down_pct = valid_group['downtimePct'].max()
            
        total_downtime = group['downtime'].sum()
        
        # KPI Target: <= 3.5%
        status = "Pass" if avg_down_pct <= 3.5 else "Fail"
        
        summary_data.append({
            'Asset Code': asset,
            'Avg Utilization%': avg_util,
            'Total Downtime (h)': total_downtime,
            'Avg Downtime%': avg_down_pct,
            'Max Downtime%': max_down_pct,
            'KPI Status': status
        })
        
    df_summary = pd.DataFrame(summary_data)
    
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Add sheets
    ws_daily = wb.create_sheet(title="Daily Detail")
    for r in dataframe_to_rows(df_daily, index=False, header=True):
        ws_daily.append(r)
        
    ws_summary = wb.create_sheet(title="Weekly Summary")
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)
        
    # Define styles
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def apply_styles(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "B2"
        
    apply_styles(ws_daily)
    apply_styles(ws_summary)
    
    # Conditional formatting - Daily Detail
    for row in ws_daily.iter_rows(min_row=2, min_col=1, max_col=7):
        dt_cell = row[5] # F: Downtime %
        idle_cell = row[6] # G: Idle %
        
        if isinstance(dt_cell.value, (int, float)):
            dt_cell.number_format = '0.00"%"'
            if dt_cell.value > 3.5:
                dt_cell.fill = red_fill
                dt_cell.font = Font(color="FFFFFF", bold=True)
                
        if isinstance(idle_cell.value, (int, float)):
            idle_cell.number_format = '0.00"%"'
            if idle_cell.value > 10.3:
                idle_cell.fill = yellow_fill
                
    # Conditional formatting - Summary
    for row in ws_summary.iter_rows(min_row=2, min_col=1, max_col=6):
        for col_idx in [1, 3, 4]: 
            cell = row[col_idx]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00"%"'
                
        cell_dt = row[2]
        if isinstance(cell_dt.value, (int, float)):
            cell_dt.number_format = '0.00'
            
    wb.save(filename)
